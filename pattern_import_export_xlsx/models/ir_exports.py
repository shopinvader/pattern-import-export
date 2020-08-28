# Copyright 2020 Akretion France (http://www.akretion.com)
# License AGPL-3.0 or later (http://www.gnu.org/licenses/agpl).
# pylint: disable=missing-manifest-dependency
import base64
from io import BytesIO

import openpyxl
from openpyxl.utils import get_column_letter, quote_sheetname
from openpyxl.worksheet.datavalidation import DataValidation

from odoo import _, api, fields, models
from odoo.exceptions import UserError


class IrExports(models.Model):
    _inherit = "ir.exports"

    export_format = fields.Selection(selection_add=[("xlsx", "Excel")])
    tab_to_import = fields.Selection([("first", "First"), ("match_name", "Match Name")])

    @api.multi
    def _create_xlsx_file(self, records):
        self.ensure_one()
        book = openpyxl.Workbook()
        main_sheet = self._build_main_sheet_structure(book)
        self._populate_main_sheet_rows(main_sheet, records)
        tab_data = self.export_fields._get_tab_data()
        self._create_tabs(book, tab_data)
        main_sheet_length = len(records.ids) + 1
        self._create_validators(main_sheet, main_sheet_length, tab_data)
        book.close()
        xlsx_file = BytesIO()
        book.save(xlsx_file)
        return xlsx_file

    def _build_main_sheet_structure(self, book):
        """
        Write main sheet header and other style details
        """
        main_sheet = book["Sheet"]
        main_sheet.title = self.name
        for col, header in enumerate(self._get_header(), start=1):
            main_sheet.cell(row=1, column=col, value=header)
        return main_sheet

    def _populate_main_sheet_rows(self, main_sheet, records):
        """
        Get the actual data and write it row by row on the main sheet
        """
        headers = self._get_header()
        for row, values in enumerate(self._get_data_to_export(records), start=2):
            for col, header in enumerate(headers, start=1):
                main_sheet.cell(row=row, column=col, value=values.get(header, ""))

    def _create_tabs(self, book, tab_data):
        """ Create additional sheets for export lines with create tab option
        and write all valid choices """
        for name, headers, data, __ in tab_data:
            new_sheet = book.create_sheet(name)
            for col_number, header in enumerate(headers, start=1):
                new_sheet.cell(row=1, column=col_number, value=header)
            for row_number, row_data in enumerate(data, start=2):
                for col_number, cell_data in enumerate(row_data, start=1):
                    new_sheet.cell(row=row_number, column=col_number, value=cell_data)

    def _create_validators(self, main_sheet, main_sheet_length, tab_data):
        """ Add validators: source permitted records from tab sheets,
        apply validation to main sheet """
        for el in tab_data:
            tab_name, _, data, col_dst = el
            col_letter_dst = get_column_letter(col_dst)
            # TODO support arbitrary columns/attributes instead of
            #  only name
            col_letter_src = get_column_letter(1)
            range_src = "${}$2:${}${}".format(
                col_letter_src, col_letter_src, str(1 + len(data))
            )
            formula_range_src = "=" + quote_sheetname(tab_name) + "!" + range_src
            validation = DataValidation(type="list", formula1=formula_range_src)
            range_dst = "${}$2:${}${}".format(
                col_letter_dst, col_letter_dst, str(main_sheet_length)
            )
            validation.add(range_dst)
            main_sheet.add_data_validation(validation)

    @api.multi
    def _export_with_record_xlsx(self, records):
        """
        Export given recordset
        @param records: recordset
        @return: string
        """
        self.ensure_one()
        excel_file = self._create_xlsx_file(records)
        return excel_file.getvalue()

    # Import part

    def _get_worksheet(self, workbook):
        name = None
        if self.tab_to_import == "first":
            name = workbook.sheetnames[0]
        elif self.tab_to_import == "match_name":
            for sheetname in workbook.sheetnames:
                if sheetname.lower() == self.name.lower():
                    name = sheetname
                    break
            if not name:
                raise UserError(
                    _("The file do not contain tab with the name {}").format(self.name)
                )
        else:
            raise UserError(_("Please select a tab to import on the pattern"))
        return workbook[name]

    def _find_real_last_column(self, worksheet):
        """
        The last column and row are actually written in the excel file
        Openpyxl doesn't automatically verify if it is right or not
        """
        tentative_last_column = worksheet.max_column
        for col in reversed(range(tentative_last_column)):
            if worksheet.cell(1, col + 1).value:
                break
        return col + 1

    def _find_real_last_row(self, worksheet, max_col):
        """ See _find_real_last_column """
        tentative_last_row = worksheet.max_row
        for row in reversed(range(tentative_last_row)):
            row_has_val = any(
                worksheet.cell(row + 1, col + 1).value for col in range(max_col)
            )
            if row_has_val:
                break
        return row + 1

    @api.multi
    def _read_import_data_xlsx(self, datafile):
        # note that columns and rows are 1-based
        workbook = openpyxl.load_workbook(BytesIO(datafile), data_only=True)
        worksheet = self._get_worksheet(workbook)
        headers = []
        real_last_column = self._find_real_last_column(worksheet)
        for col in range(real_last_column):
            headers.append(worksheet.cell(1, col + 1).value)
        real_last_row = self._find_real_last_row(worksheet, real_last_column)
        for row in range(real_last_row - 1):
            elm = {}
            for col in range(real_last_column):
                elm[headers[col]] = worksheet.cell(row + 2, col + 1).value
            yield elm

    def _process_load_result_for_xls(self, attachment, res):
        infile = BytesIO(base64.b64decode(attachment.datas))
        wb = openpyxl.load_workbook(filename=infile)
        ws = self._get_worksheet(wb)
        global_message = []
        if ws["A1"].value != _("#Error"):
            ws.insert_cols(1)
            ws.cell(1, 1, value=_("#Error"))
        for message in res["messages"]:
            if "rows" in message:
                ws.cell(message["rows"]["to"] + 1, 1, value=message["message"].strip())
            else:
                global_message.append(message)
        output = BytesIO()
        wb.save(output)
        attachment.datas = base64.b64encode(output.getvalue())
        ids = res["ids"] or []
        info = _(
            "Number of record imported {} Number of error/warning {}"
            "\nrecord ids details: {}"
            "\n{}"
        ).format(
            len(ids),
            len(res.get("messages", [])),
            ids,
            "\n".join(
                [
                    "{}: {}".format(message["type"], message["message"])
                    for message in global_message
                ]
            ),
        )
        if res.get("messages"):
            status = "fail"
        else:
            status = "success"
        return info, status

    def _process_load_result(self, attachment, res):
        if self.export_format == "xlsx":
            return self._process_load_result_for_xls(attachment, res)
        else:
            return super()._process_load_result(attachment, res)
