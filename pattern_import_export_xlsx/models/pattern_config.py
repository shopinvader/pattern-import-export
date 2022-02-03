# Copyright 2020 Akretion France (http://www.akretion.com)
# License AGPL-3.0 or later (http://www.gnu.org/licenses/agpl).
# pylint: disable=missing-manifest-dependency
from io import BytesIO

import openpyxl
from openpyxl.utils import get_column_letter, quote_sheetname
from openpyxl.worksheet.datavalidation import DataValidation

from odoo import fields, models

EXTRA_LINE_NUMBER = 1000


class PatternConfig(models.Model):
    _inherit = "pattern.config"

    export_format = fields.Selection(selection_add=[("xlsx", "Excel")])
    tab_to_import = fields.Selection(
        [("first", "First"), ("match_name", "Match Name")], default="first"
    )

    # TODO we should move this code in pattern.file
    def _create_xlsx_file(self, records):
        self.ensure_one()
        book = openpyxl.Workbook()
        main_sheet = self._build_main_sheet_structure(book)
        self._populate_main_sheet_rows(main_sheet, records)
        tabs = self._get_metadata()["tabs"]
        self._create_tabs(book, tabs)
        self._create_validators(main_sheet, records, tabs)
        book.close()
        xlsx_file = BytesIO()
        book.save(xlsx_file)
        return xlsx_file

    def _build_main_sheet_structure(self, book):
        """
        Write main sheet header and other style details
        """
        main_sheet = book["Sheet"]
        main_sheet.title = self.name
        for row, lines in enumerate(self._get_output_headers(), start=1):
            for col, header in enumerate(lines.values(), start=1):
                main_sheet.cell(row=row, column=col, value=header)
        return main_sheet

    def _populate_main_sheet_rows(self, main_sheet, records):
        """
        Get the actual data and write it row by row on the main sheet
        """
        headers = self._get_header()
        for row, values in enumerate(
            self._get_data_to_export(records), start=self.row_start_records
        ):
            for col, header in enumerate(headers, start=1):
                main_sheet.cell(row=row, column=col, value=values.get(header, ""))

    def _create_tabs(self, book, tabs):
        """Create additional sheets for export lines with create tab option
        and write all valid choices"""
        for tab_name, tab in tabs.items():
            new_sheet = book.create_sheet(tab_name)
            for col_number, header in enumerate(tab["headers"], start=1):
                new_sheet.cell(row=1, column=col_number, value=header)
            for row_number, row_data in enumerate(tab["data"], start=2):
                for col_number, cell_data in enumerate(row_data, start=1):
                    new_sheet.cell(row=row_number, column=col_number, value=cell_data)

    def _create_validators(self, main_sheet, records, tabs):
        """Add validators: source permitted records from tab sheets,
        apply validation to main sheet"""
        main_sheet_length = len(records.ids) + EXTRA_LINE_NUMBER
        for tab_name, tab in tabs.items():
            # TODO support arbitrary columns/attributes instead of
            #  only name
            col_letter_src = get_column_letter(1)
            range_src = "${}$2:${}${}".format(
                col_letter_src,
                col_letter_src,
                str(EXTRA_LINE_NUMBER + len(tab["data"])),
            )
            formula_range_src = "=" + quote_sheetname(tab_name) + "!" + range_src
            validation = DataValidation(type="list", formula1=formula_range_src)
            for idx_col in tab["idx_col_validator"]:
                col_letter_dst = get_column_letter(idx_col)
                range_dst = "${}${}:${}${}".format(
                    col_letter_dst,
                    str(self.row_start_records),
                    col_letter_dst,
                    str(max(main_sheet_length, 2)),
                )
                validation.add(range_dst)
            main_sheet.add_data_validation(validation)

    def _export_with_record_xlsx(self, records):
        """
        Export given recordset
        @param records: recordset
        @return: string
        """
        self.ensure_one()
        excel_file = self._create_xlsx_file(records)
        return excel_file.getvalue()
