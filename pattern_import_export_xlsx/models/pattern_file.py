# Copyright 2020 Akretion (https://www.akretion.com).
# @author Sébastien BEAU <sebastien.beau@akretion.com>
# License AGPL-3.0 or later (https://www.gnu.org/licenses/agpl).

import base64
from io import BytesIO

import openpyxl

from odoo import _, models
from odoo.exceptions import UserError

STOP_AFTER_NBR_EMPTY = 10


class PatternFile(models.Model):
    _inherit = "pattern.file"

    def _get_worksheet(self, workbook):
        name = None
        tab_to_import = self.pattern_config_id.tab_to_import
        if tab_to_import == "first":
            name = workbook.sheetnames[0]
        elif tab_to_import == "match_name":
            for sheetname in workbook.sheetnames:
                if (
                    sheetname.lower().strip()
                    == self.pattern_config_id.name.lower().strip()
                ):
                    name = sheetname
                    break
            if not name:
                raise UserError(
                    _("The file do not contain tab with the name {}").format(
                        self.pattern_config_id.name
                    )
                )
        else:
            raise UserError(_("Please select a tab to import on the pattern"))
        return workbook[name]

    def _parse_data_xlsx(self, data):
        workbook = openpyxl.load_workbook(BytesIO(data), data_only=True, read_only=True)
        worksheet = self._get_worksheet(workbook)
        headers = None
        count_empty = 0
        for idx, row in enumerate(worksheet.rows):
            if self.pattern_config_id.nr_of_header_rows == idx + 1:
                headers = [x.value for x in row]
            elif headers:
                vals = [x.value for x in row]
                if any(vals):
                    count_empty = 0
                    item = dict(zip(headers, vals))
                    # we remove column without header
                    item.pop(None, "")
                    # the position return is the row number
                    # libreoffice/excel/human start from 1
                    yield idx + 1, item
                else:
                    count_empty += 1
            if count_empty > STOP_AFTER_NBR_EMPTY:
                break
        workbook.close()

    def write_error_in_xlsx(self):
        # TODO writing in an existing big excel file is long with openpyxl
        # maybe we should try some other tools
        # https://editpyxl.readthedocs.io
        infile = BytesIO(base64.b64decode(self.datas))
        wb = openpyxl.load_workbook(filename=infile)
        ws = self._get_worksheet(wb)
        ws.column_dimensions["A"].hidden = False

        # we clear the error col if exist
        if ws["A1"].value == _("#Error"):
            ws.delete_cols(1)
        ws.insert_cols(1)
        ws.cell(1, 1, value=_("#Error"))
        for chunk in self.chunk_ids.filtered(lambda c: c.nbr_error > 0):
            for message in chunk.messages:
                if "rows" in message:
                    ws.cell(message["rows"]["from"], 1, value=message["message"])
                else:
                    # If no row are specify, this is a global message
                    # that should be applied until the end of the chunk
                    for idx in range(chunk.start_idx, chunk.stop_idx + 1):
                        if self.pattern_config_id.nr_of_header_rows == 2 and idx == 2:
                            idx = idx + 1
                        ws.cell(idx, 1, value=message["message"])
        output = BytesIO()
        wb.save(output)
        self.datas = base64.b64encode(output.getvalue())

    def set_import_done(self):
        super().set_import_done()
        for record in self:
            if (
                record.state == "failed"
                and record.pattern_config_id.export_format == "xlsx"
            ):
                record.write_error_in_xlsx()
        return True
