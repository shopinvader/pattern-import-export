# Copyright 2022 Akretion (https://www.akretion.com).
# @author Sébastien BEAU <sebastien.beau@akretion.com>
# License AGPL-3.0 or later (https://www.gnu.org/licenses/agpl).


import base64
from io import BytesIO

import openpyxl

from odoo.tests.common import SavepointCase

from odoo.addons.pattern_import_export.tests.common import PatternCommon


class CommonPatternExportExcel(PatternCommon, SavepointCase):
    @classmethod
    def _set_up_tab_names(cls):
        for el in ("ignore_one", "countries_1", "countries_2"):
            attr_name_tab = "tab_name_" + el
            attr_name_filter = "filter_" + el
            filter_id = getattr(cls, attr_name_filter)
            attr_val_tab = "({}) {}".format(filter_id.id, filter_id.name)
            setattr(cls, attr_name_tab, attr_val_tab)

    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        cls._set_up_tab_names()
        for el in cls.pattern_config, cls.pattern_config_m2m, cls.pattern_config_o2m:
            el.export_format = "xlsx"

    def _read_worksheet(self, pattern_config, records=None, sheet_name=None):
        if not records:
            # we only export the header
            records = self.env[pattern_config.resource]
        wb = self._helper_get_resulting_wb(pattern_config, records)
        if not sheet_name:
            sheet = wb.worksheets[0]
        else:
            sheet = wb[sheet_name]
        for idx, row in enumerate(sheet.rows):
            yield idx, row

    def _get_header(self, pattern_config, use_description=False, sheet_name=None):
        if use_description:
            self.pattern_config.header_format = "description_and_tech"
        for _idx, row in self._read_worksheet(pattern_config, sheet_name=sheet_name):
            return [x.value for x in row]

    def _get_data(self, pattern_config, records=None, sheet_name=None):
        data = []
        headers = None
        for idx, row in self._read_worksheet(
            pattern_config, records=records, sheet_name=sheet_name
        ):
            vals = [x.value for x in row]
            if idx == 0:
                headers = vals
            elif any(vals):
                data.append(dict(zip(headers, vals)))
            else:
                break
        return data

    def _helper_get_resulting_wb(self, export, records):
        export._export_with_record(records)
        attachment = self._get_attachment(export)
        self.assertEqual(attachment.name, export.name + ".xlsx")
        decoded_data = base64.b64decode(attachment.datas)
        decoded_obj = BytesIO(decoded_data)
        return openpyxl.load_workbook(decoded_obj)

    def _helper_check_cell_values(self, sheet, expected_values):
        """To allow for csv-like syntax in tests, just give a list
        of lists, with 1 list <=> 1 row"""
        for itr_row, row in enumerate(expected_values, start=2):
            for itr_col, expected_cell_value in enumerate(row, start=1):
                cell_value = sheet.cell(row=itr_row, column=itr_col).value
                self.assertEqual(cell_value, expected_cell_value)

    def _helper_check_headers(self, sheet, expected_headers):
        for itr_col, expected_cell_value in enumerate(expected_headers, start=1):
            cell_value = sheet.cell(row=1, column=itr_col).value
            self.assertEqual(cell_value, expected_cell_value)
