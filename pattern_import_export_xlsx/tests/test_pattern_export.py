# Copyright 2020 Akretion France (http://www.akretion.com)
# License AGPL-3.0 or later (http://www.gnu.org/licenses/agpl).
# pylint: disable=missing-manifest-dependency

import base64
from io import BytesIO

import openpyxl

from odoo.tests.common import SavepointCase

# pylint: disable=odoo-addons-relative-import
from odoo.addons.pattern_import_export.tests.common import PatternCommon
from odoo.addons.pattern_import_export.tests.test_pattern_export import (
    PatternCaseExport,
)

CELL_VALUE_EMPTY = None


class TestPatternExportExcel(PatternCaseExport, PatternCommon, SavepointCase):
    @classmethod
    def _set_up_tab_names(cls):
        for el in ("ignore_one", "countries_1", "countries_2"):
            attr_name_tab = "tab_name_" + el
            attr_name_filter = "filter_" + el
            filter_id = getattr(cls, attr_name_filter)
            attr_val_tab = "({}) {}".format(filter_id.id, filter_id.name)
            setattr(cls, attr_name_tab, attr_val_tab)

    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        cls._set_up_tab_names()
        for el in cls.pattern_config, cls.pattern_config_m2m, cls.pattern_config_o2m:
            el.export_format = "xlsx"

    def _read_worksheet(self, pattern_config, records=None, sheet_name=None):
        if not records:
            # we only export the header
            records = self.env[pattern_config.resource]
        wb = self._helper_get_resulting_wb(pattern_config, records)
        if not sheet_name:
            sheet = wb.worksheets[0]
        else:
            sheet = wb[sheet_name]
        for idx, row in enumerate(sheet.rows):
            yield idx, row

    def _get_header(self, pattern_config, use_description=False, sheet_name=None):
        if use_description is not None:
            self.pattern_config.use_description = use_description
        for _idx, row in self._read_worksheet(pattern_config, sheet_name=sheet_name):
            return [x.value for x in row]

    def _get_data(self, pattern_config, records=None, sheet_name=None):
        data = []
        headers = None
        for idx, row in self._read_worksheet(
            pattern_config, records=records, sheet_name=sheet_name
        ):
            vals = [x.value for x in row]
            if idx == 0:
                headers = vals
            elif any(vals):
                data.append(dict(zip(headers, vals)))
            else:
                break
        return data

    def _helper_get_resulting_wb(self, export, records):
        export._export_with_record(records)
        attachment = self._get_attachment(export)
        self.assertEqual(attachment.name, export.name + ".xlsx")
        decoded_data = base64.b64decode(attachment.datas)
        decoded_obj = BytesIO(decoded_data)
        return openpyxl.load_workbook(decoded_obj)

    def _helper_check_cell_values(self, sheet, expected_values):
        """To allow for csv-like syntax in tests, just give a list
        of lists, with 1 list <=> 1 row"""
        for itr_row, row in enumerate(expected_values, start=2):
            for itr_col, expected_cell_value in enumerate(row, start=1):
                cell_value = sheet.cell(row=itr_row, column=itr_col).value
                self.assertEqual(cell_value, expected_cell_value)

    def _helper_check_headers(self, sheet, expected_headers):
        for itr_col, expected_cell_value in enumerate(expected_headers, start=1):
            cell_value = sheet.cell(row=1, column=itr_col).value
            self.assertEqual(cell_value, expected_cell_value)

    def test_export_tabs(self):
        wb = self._helper_get_resulting_wb(self.pattern_config, self.partners)
        sheets_name = [x.title for x in wb.worksheets]
        sheet_tab_2 = wb[self.tab_name_countries_1]
        sheet_tab_3 = wb["Tags"]
        self.assertEqual(sheets_name, ["Partner", self.tab_name_countries_1, "Tags"])
        expected_values_tab_2 = [["BE"], ["FR"], ["US"], [CELL_VALUE_EMPTY]]
        expected_values_tab_3 = [
            ["Consulting Services"],
            ["Desk Manufacturers"],
            ["Employees"],
            ["Office Supplies"],
            ["Prospects"],
            ["Services"],
            ["Vendor"],
            [CELL_VALUE_EMPTY],
        ]
        self._helper_check_cell_values(sheet_tab_2, expected_values_tab_2)
        self._helper_check_cell_values(sheet_tab_3, expected_values_tab_3)

    def test_export_tabs_name_no_filter(self):
        self.pattern_config.export_fields[3].tab_filter_id = self.env["ir.filters"]
        wb = self._helper_get_resulting_wb(self.pattern_config, self.partners)
        expected_tab_name = "Country"
        self.assertTrue(wb[expected_tab_name])

    def test_export_tabs_name_long_name(self):
        # 32 - 4 = 28 characters, plus "(id) " = 32 characters, the limit
        self.pattern_config.export_fields[
            3
        ].tab_filter_id.name = "CIOIOIOIOIOIOIOIOIOIOIOIOIOI"
        wb = self._helper_get_resulting_wb(self.pattern_config, self.partners)
        expected_tab_name = "({}) " "CIOIOIOIOIOIOIOIOIOIOIOI...".format(
            self.filter_countries_1.id
        )
        self.assertTrue(wb[expected_tab_name])

    def test_export_tabs_with_subpattern(self):
        wb = self._helper_get_resulting_wb(self.pattern_config_o2m, self.partners)
        sheets_name = [x.title for x in wb.worksheets]
        self.assertEqual(len(sheets_name), 3)
        sheet_tab_2 = wb[self.tab_name_countries_1]
        sheet_tab_3 = wb["Tags"]
        self.assertEqual(
            sheets_name, ["Partner with contact", self.tab_name_countries_1, "Tags"]
        )
        expected_values_tab_2 = [["BE"], ["FR"], ["US"], [CELL_VALUE_EMPTY]]
        expected_values_tab_3 = [
            ["Consulting Services"],
            ["Desk Manufacturers"],
            ["Employees"],
            ["Office Supplies"],
            ["Prospects"],
            ["Services"],
            ["Vendor"],
            [CELL_VALUE_EMPTY],
        ]
        self._helper_check_cell_values(sheet_tab_2, expected_values_tab_2)
        self._helper_check_cell_values(sheet_tab_3, expected_values_tab_3)

    def test_export_validators_simple(self):
        wb = self._helper_get_resulting_wb(self.pattern_config, self.partners)
        sheet_base = wb["Partner"]
        self.assertEqual(
            sheet_base.data_validations.dataValidation[0].formula1,
            "='{}'!$A$2:$A$4".format(self.tab_name_countries_1),
        )
        self.assertEqual(
            str(sheet_base.data_validations.dataValidation[0].cells), "D2:D1000"
        )
        self.assertEqual(
            sheet_base.data_validations.dataValidation[1].formula1,
            "='Tags'!$A$2:$A$8",
        )
        self.assertEqual(
            str(sheet_base.data_validations.dataValidation[1].cells), "E2:E1000"
        )

    def test_export_validators_simple_with_subpattern(self):
        wb = self._helper_get_resulting_wb(self.pattern_config_o2m, self.partners)
        sheet_base = wb["Partner with contact"]
        self.assertEqual(
            sheet_base.data_validations.dataValidation[0].formula1,
            "='{}'!$A$2:$A$4".format(self.tab_name_countries_1),
        )
        self.assertEqual(
            str(sheet_base.data_validations.dataValidation[0].cells),
            "F2:F1000 K2:K1000 P2:P1000 R2:R1000",
        )
        self.assertEqual(
            sheet_base.data_validations.dataValidation[1].formula1,
            "='Tags'!$A$2:$A$8",
        )
        self.assertEqual(
            str(sheet_base.data_validations.dataValidation[1].cells),
            "G2:G1000 L2:L1000 Q2:Q1000",
        )

    def test_export_validators_many2many(self):
        self.pattern_config_m2m.export_fields[2].number_occurence = 3
        wb = self._helper_get_resulting_wb(self.pattern_config_m2m, self.users)
        sheet_base = wb["Users list - M2M"]
        self.assertEqual(
            sheet_base.data_validations.dataValidation[0].formula1,
            "='{}'!$A$2:$A$4".format(self.tab_name_ignore_one),
        )
        for idx, col_letter in enumerate(("C", "D", "E")):
            self.assertEqual(
                str(sheet_base.data_validations.dataValidation[0].cells.ranges[idx]),
                "{0}2:{0}1000".format(col_letter),
            )

    def test_export_m2m_tabs(self):
        wb = self._helper_get_resulting_wb(self.pattern_config_m2m, self.users)
        sheet_tab_2 = wb[self.tab_name_ignore_one]
        expected_values_tab_2 = [
            ["Awesome company"],
            ["Bad company"],
            ["YourCompany"],
            [CELL_VALUE_EMPTY],
        ]
        self._helper_check_cell_values(sheet_tab_2, expected_values_tab_2)

    def test_export_m2m_validators(self):
        wb = self._helper_get_resulting_wb(self.pattern_config_m2m, self.users)
        sheet_base = wb["Users list - M2M"]
        self.assertEqual(
            sheet_base.data_validations.dataValidation[0].formula1,
            "='{}'!$A$2:$A$4".format(self.tab_name_ignore_one),
        )
        self.assertEqual(
            str(sheet_base.data_validations.dataValidation[0].cells), "C2:C1000"
        )
