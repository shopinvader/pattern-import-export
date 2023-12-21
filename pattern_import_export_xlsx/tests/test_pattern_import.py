# Copyright 2020 Akretion (https://www.akretion.com).
# @author Sébastien BEAU <sebastien.beau@akretion.com>
# License AGPL-3.0 or later (https://www.gnu.org/licenses/agpl).
import base64
from io import BytesIO
from os import path

# TODO FIXME somehow Travis complains that openpyxl isn't there,
# the warning shows only here and not in any other import of openpyxl?
# pylint: disable=missing-manifest-dependency
import openpyxl
from mock import Mock

from odoo.tests import SavepointCase
from odoo.tools import mute_logger

# helper to dump the result of the import into an excel file
DUMP_OUTPUT = False


PATH = path.dirname(__file__) + "/fixtures/"


class TestPatternImportExcel(SavepointCase):
    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        cls.env = cls.env(
            context=dict(
                cls.env.context,
                test_queue_job_no_delay=True,  # no jobs thanks
            )
        )
        cls.pattern_config_partner = cls.env["pattern.config"].create(
            {"name": "Partner", "resource": "res.partner", "export_format": "xlsx"}
        )
        cls.pattern_config_users = cls.env["pattern.config"].create(
            {"name": "User", "resource": "res.users", "export_format": "xlsx"}
        )
        cls.user_admin = cls.env.ref("base.user_admin")
        cls.user_demo = cls.env.ref("base.user_demo")
        cls.env.cr.commit = Mock()

    @classmethod
    def _load_file(cls, filename, pattern_config_id):
        data = base64.b64encode(open(PATH + filename, "rb").read())
        wizard = cls.env["import.pattern.wizard"].create(
            {
                "pattern_config_id": pattern_config_id.id,
                "import_file": data,
                "filename": "example.xlsx",
            }
        )
        pattern_file = wizard.action_launch_import()

        if DUMP_OUTPUT:
            attachment = cls.env["pattern.file"].search([], limit=1, order="id desc")
            output_name = filename.replace(".xlsx", ".result.xlsx")
            with open(output_name, "wb") as output:
                output.write(base64.b64decode(attachment.datas))
        return pattern_file

    def test_import_partners_ok(self):
        """
        * Lookup by email
        * Update some o2m fields
        """
        self._load_file("example.partners.ok.xlsx", self.pattern_config_partner)
        # check first line
        partner = self.env.ref("base.res_partner_1")

        self.assertEqual(partner.name, "Wood Corner Updated")
        self.assertEqual(partner.email, "wood.corner26@example.com")
        self.assertEqual(partner.phone, "111111111")
        self.assertEqual(partner.country_id.code, "FR")

        # In demo data we have 3 childs
        # after the import we should still have 3 childs
        # In a long term we should have an option
        # to active/remove unwanted o2m
        self.assertEqual(len(partner.child_ids), 3)

        contact_1, contact_2, contact_3 = partner.child_ids.sorted("id")
        self.assertEqual(contact_1.id, self.env.ref("base.res_partner_address_1").id)
        self.assertEqual(contact_1.name, "Willie Burke Updated")
        self.assertEqual(contact_1.email, "willie.burke80.updated@example.com")
        self.assertEqual(contact_1.function, "Service Manager")

        self.assertEqual(contact_2.id, self.env.ref("base.res_partner_address_2").id)
        self.assertEqual(contact_2.name, "Ron Gibson Updated")
        self.assertEqual(contact_2.email, "ron.gibson76.updated@example.com")
        self.assertEqual(contact_2.function, "Store Manager")

        # check second line
        partner = self.env.ref("base.res_partner_2")

        self.assertEqual(partner.name, "Deco Addict Updated")
        self.assertEqual(partner.country_id.code, "DE")

        self.assertEqual(len(partner.child_ids), 3)

        contact_1, contact_2, contact_3 = partner.child_ids.sorted("id")
        self.assertEqual(contact_1.id, self.env.ref("base.res_partner_address_3").id)
        self.assertEqual(contact_2.id, self.env.ref("base.res_partner_address_4").id)

        # check three line creation

        partner = self.env["res.partner"].search(
            [("email", "=", "akretion-pattern@example.com")]
        )
        self.assertEqual(len(partner), 1)
        self.assertEqual(partner.name, "Akretion")
        self.assertEqual(partner.phone, "333333333")
        self.assertEqual(partner.country_id.code, "FR")
        self.assertEqual(len(partner.child_ids), 2)

        contact_1, contact_2 = partner.child_ids.sorted("id")
        self.assertEqual(contact_1.name, "Sebastien")
        self.assertEqual(contact_1.email, "seb-pattern@example.com")
        self.assertEqual(contact_1.function, "Service Manager")

        self.assertEqual(contact_2.name, "Raph")
        self.assertEqual(contact_2.email, "raph-pattern@example.com")
        self.assertEqual(contact_2.function, "Store Manager")

    @mute_logger("odoo.sql_db")
    def test_partial_import(self):
        pattern_file = self._load_file(
            "example.partners.failed.xlsx", self.pattern_config_partner
        )
        self.env.clear()

        self.assertEqual(pattern_file.state, "failed")

        # Check that the data without error are updated
        partner = self.env.ref("base.res_partner_1")
        self.assertEqual(partner.name, "Wood Corner Updated")

        partner = self.env.ref("base.res_partner_2")
        self.assertEqual(partner.name, "Deco Addict Updated")

        partner = self.env["res.partner"].search(
            [("email", "=", "akretion-pattern@example.com")]
        )
        self.assertEqual(len(partner), 1)

        infile = BytesIO(base64.b64decode(pattern_file.datas))
        wb = openpyxl.load_workbook(filename=infile)
        ws = wb.worksheets[0]
        self.assertEqual(ws["A1"].value, "#Error")
        self.assertIsNone(ws["A2"].value)
        self.assertIsNone(ws["A3"].value)
        self.assertIsNone(ws["A4"].value)
        self.assertEqual("'Contacts require a name'", ws["A5"].value)

    @mute_logger("odoo.sql_db")
    def test_partial_import_too_many_error(self):
        pattern_file = self._load_file(
            "example.partners.too_many_error.xlsx", self.pattern_config_partner
        )
        self.env.clear()

        self.assertEqual(pattern_file.state, "failed")
        # Check that the data without error are updated
        partner = self.env.ref("base.res_partner_1")
        self.assertEqual(partner.name, "Wood Corner Updated")

        partner = self.env.ref("base.res_partner_2")
        self.assertEqual(partner.name, "Deco Addict Updated")

        # Last line should not have been imported as
        # odoo break when there is too many error
        partner = self.env["res.partner"].search(
            [("email", "=", "akretion-pattern@example.com")]
        )
        self.assertEqual(len(partner), 0)

        infile = BytesIO(base64.b64decode(pattern_file.datas))
        wb = openpyxl.load_workbook(filename=infile)
        ws = wb.worksheets[0]
        self.assertEqual(ws["A1"].value, "#Error")
        self.assertIsNone(ws["A2"].value)
        self.assertIsNone(ws["A3"].value)

        for idx in range(4, 13):
            self.assertEqual("'Contacts require a name'", ws["A%s" % idx].value)
        for idx in range(13, 21):
            self.assertEqual(
                "Found more than 10 errors and more than one error per 10 records, "
                "interrupted to avoid showing too many errors.",
                ws["A%s" % idx].value,
            )
        self.assertIsNone(ws["A21"].value)

    @mute_logger("odoo.sql_db")
    def test_multi_chunk(self):
        self.pattern_config_partner.chunk_size = 5
        self.pattern_config_partner.process_multi = True
        pattern_file = self._load_file(
            "example.partners.too_many_error.xlsx", self.pattern_config_partner
        )
        self.env.clear()

        self.assertEqual(pattern_file.state, "failed")
        # Check that the data without error are updated
        partner = self.env.ref("base.res_partner_1")
        self.assertEqual(partner.name, "Wood Corner Updated")

        partner = self.env.ref("base.res_partner_2")
        self.assertEqual(partner.name, "Deco Addict Updated")

        # Last line should have been imported as
        # we have chunk of 5 so we can not have 10 error
        partner = self.env["res.partner"].search(
            [("email", "=", "akretion-pattern@example.com")]
        )
        self.assertEqual(len(partner), 1)

        infile = BytesIO(base64.b64decode(pattern_file.datas))
        wb = openpyxl.load_workbook(filename=infile)
        ws = wb.worksheets[0]
        self.assertEqual(ws["A1"].value, "#Error")
        self.assertIsNone(ws["A2"].value)
        self.assertIsNone(ws["A3"].value)
        for idx in range(4, 20):
            self.assertEqual("'Contacts require a name'", ws["A%s" % idx].value)
        self.assertIsNone(ws["A20"].value)

    def test_import_users_ok(self):
        """
        * Lookup by DB ID
        * Simple update
        """
        self._load_file("example.users.ok.xlsx", self.pattern_config_users)
        self.assertEqual(self.user_admin.name, "Mitchell Admin Updated")
        self.assertEqual(self.user_demo.name, "Marc Demo Updated")

    def test_import_users_descriptive_ok(self):
        """
        * Use descriptive headers
        * Lookup by DB ID
        * Simple update
        """
        self.pattern_config_users.header_format = "description_and_tech"
        self._load_file("example.users.descriptive.ok.xlsx", self.pattern_config_users)
        self.assertEqual(self.user_admin.name, "Mitchell Admin Updated")
        self.assertEqual(self.user_demo.name, "Marc Demo Updated")

    # TODO FIXME
    @mute_logger("odoo.sql_db")
    def disable_test_import_users_failed(self):
        """
        * Lookup by external ID
        * Report error in excel file through external id not found
        """
        self._load_file("example.users.failed.xlsx", self.pattern_config_users)
        attachment = self.env["ir.attachment"].search([], order="id desc", limit=1)
        infile = BytesIO(base64.b64decode(attachment.datas))
        wb = openpyxl.load_workbook(filename=infile)
        ws = wb.worksheets[0]
        self.assertEqual(ws["A1"].value, "#Error")
        self.assertTrue(ws["A2"].value)
        self.assertTrue(ws["A3"].value)

    def test_import_partners_with_parents(self):
        self._load_file("example.partners.parent.xlsx", self.pattern_config_partner)
        partner_parent = self.env["res.partner"].search([("name", "=", "Apple")])
        self.assertTrue(partner_parent)
        partner_child = self.env["res.partner"].search([("name", "=", "Steve Jobs")])
        self.assertTrue(partner_child.parent_id == partner_parent)
